# -*- coding:utf-8 -*-

from fetch import fetch, parse, fetch2, fetch3, fetch4, fetch5
import xlwt
import time
from config import sheetHeader, sheetName
from openpyxl import Workbook


def output(reports, target):
    results = {}
    if reports == "ALL":
        reports = "1000,1001,1002,1003,1004,1005,1006,1007,1008,1009,1010,1012,1013,1014,1020"
    templates = reports.split(',')
    for template in templates:
        if template != '1020':
            try:
                data = parse('template/%s.xml' % template)
            except Exception, e:
                print e
                print 'template:', template
                exit(1)
        if template == '1007' or template == '1008':
            results[template] = fetch2(data['sqls'], target)
        elif template == '1011':
            results[template] = fetch4(data['sqls'])
        elif template == '1015':
            results[template] = fetch5(data['sqls'], target)
        elif template == '1020':
            fetch3_result = fetch3()
            results['1020'] = fetch3_result[1]
            sheetHeader['1020'] = fetch3_result[0]
        else:
            results[template] = fetch(data['sqls'], target)
    return results


def write(results):
    file = xlwt.Workbook()
    for k, v in results.items():
        table = file.add_sheet(sheetName[k])
        r = 1
        c = 0
        for header in sheetHeader[k]:
            table.write(0, c, header)
            c += 1
        for row in v:
            c = 0
            for cell in row:
                table.write(r, c, cell)
                c += 1
            r += 1
    file.save('report/DAT_report_%s.xls' %
              time.strftime('%Y%m%d%H%M', time.localtime()))


def writeXLSX(results, target):
    wb = Workbook()
    for k, v in results.items():
        ws = wb.create_sheet(sheetName[k])
        r = 2
        c = 1
        for header in sheetHeader[k]:
            ws.cell(row=1, column=c).value = header
            # table.write(0, c, header)
            c += 1
        for row in v:
            c = 1
            for cell_content in row:
                ws.cell(row=r, column=c).value = cell_content
                # table.write(r, c, cell)
                c += 1
            r += 1
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save('report/DAT_report_%s.xlsx' % target)
